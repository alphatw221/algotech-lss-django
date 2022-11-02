import os
import config
import django
try:
    os.environ['DJANGO_SETTINGS_MODULE'] = config.DJANGO_SETTINGS
    django.setup()
except Exception as e:
    pass

from django.conf import settings
from api import models

import traceback
import database
import service
from openpyxl import load_workbook
from io import BytesIO
import lib
from datetime import datetime
import business_policy

subscription_plan_map = {'TRI':'trial','LITE':'lite','PRM':'premium','STD':'standard'}

def imoprt_account_job(file, room_id):
    try:
        # workbook = load_workbook(filename=BytesIO(file.read()))
        workbook = load_workbook(filename=file)

        worksheet = workbook.worksheets[0]

        for row in worksheet.iter_rows(min_row=504, max_row=worksheet.max_row):
            

            try:
                cols = list(row)
                print(cols[0].coordinate)


                email = cols[0].value
                username = email.split('@')[0]
                password = str(cols[1].value)
                
                application, subscription_plan_name, unit, month_name, year = str(cols[2].value).split(' ')

                unit = int(unit)
                month = datetime.strptime(month_name, '%b').month
                signup_date = datetime(int(year), month=month, day=1)
                contactNumber = ''
                country_code = 'MY'

                purchase_price = float(cols[3].value)

                subscription_plan = subscription_plan_map.get(subscription_plan_name,'standard')
                if country_code not in business_policy.subscription_plan.SubscriptionPlan.support_country:
                    raise ''
                
                if subscription_plan not in business_policy.subscription_plan.SubscriptionPlan.support_plan:
                    raise ''
                
                # print(email)
                # print(username)
                # print(password)
                # print(signup_date)
                # print(contactNumber)
                # print(country_code)
                # print(unit)
                # print(purchase_price)
                # print(subscription_plan)

                ret = lib.helper.register_helper.create_new_account_for_james(
                    country_code=country_code, 
                    usbscription_plan=subscription_plan, 
                    username=username, 
                    email=email, 
                    password=password, 
                    signup_date=signup_date, 
                    contactNumber=contactNumber,
                    unit=unit,
                    purchase_price=purchase_price
                    )

                service.stripe.stripe.create_checkout_session_for_james(settings.STRIPE_API_KEY, 'USD', email,  application, subscription_plan_name, unit, month_name, year, purchase_price)

                # service.channels.account_import.send_success_data(room_id,ret)
            except Exception:
                print(traceback.format_exc())
                # service.channels.account_import.send_error_data(room_id,{'detail':'error'})

        # service.channels.account_import.send_complete_data(room_id,{'detail':'complete'})
    except Exception:
        print(traceback.format_exc())
        # service.channels.account_import.send_complete_data(room_id,{'detail':'error'})

  
