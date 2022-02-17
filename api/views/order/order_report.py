from api.utils.common.verify import Verify

from backend.pymongo.mongodb import db
import openpyxl
import xlsxwriter, os.path 
from api.utils.error_handle.error_handler.api_error_handler import api_error_handler

def verify_request(api_user, platform_name, platform_id, campaign_id):
    Verify.verify_user(api_user)
    platform = Verify.get_platform(api_user, platform_name, platform_id)
    campaign = Verify.get_campaign_from_platform(platform, campaign_id)

    return platform, campaign


#TODO filter by culumn , some column not print in excel 
def check_exist_report(campaign_id, column_list):
    # column_list = ['id', 'customer_name', 'platform', 'email', 'phone', 'first_name', 'last_name', 'gender', 'total', 'tax', 'currency', 'shipping_first_name', 'shipping_last_name', 'shipping_phone', 'shipping_postcode', 'shipping_region', 'shipping_location', 'shipping_address', 'shipping_method']

    campaign_title = db.api_campaign.find_one({'id': campaign_id})['title']
    file_path = '/mnt/c/Users/derek/Downloads/' + campaign_title + '_order_report.xlsx'

    # check if excel exist
    if os.path.isfile(file_path):
        existbook = openpyxl.load_workbook(file_path)
        sheet = existbook.active
        row_len = len(sheet['A'])

        # check is duplicate order_id 
        exist_file_id_list = []
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=row_len, max_col=1):
            for cell in row:
                exist_file_id_list.append(cell.value)
        campaign_order_id_list = db.api_order.find({'campaign_id': campaign_id}).distinct('id')

        if set(exist_file_id_list) == set(campaign_order_id_list):
            # if same downlaod exist file
            print ('same file')
            generate_order_report(campaign_id, column_list, file_path)
        else:
            generate_order_report(campaign_id, column_list, file_path)
    else:
        generate_order_report(campaign_id, column_list, file_path)


def generate_order_report(campaign_id, column_list, file_path):
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()
    header = workbook.add_format({
        'bold': True,
        'bg_color': '#F7F7F7',
        'color': 'black',
        'align': 'center',
        'valign': 'top',
        'border': 1
    })
    int_center = workbook.add_format({
        'align': 'center'
    })

    row, column = 0, 0
    for column_title in column_list:
        worksheet.write(row, column, column_title, header)
        column += 1
    row += 1
    column = 0

    order_id_list = []
    order_ids = db.api_order.find({'campaign_id': campaign_id})
    for order_id in order_ids:
        order_id_list.append(order_id['id'])
    
    # for each order data 1 by 1
    for order_id in order_id_list:
        order_data = db.api_order.find_one({'id': order_id})

        for column_title in column_list:
            data = order_data[column_title]
            if type(data) == int or type(data) == float:
                worksheet.write(row, column, data, int_center)
            else:
                worksheet.write(row, column, data)

            # if len(str(data)) > 8:
            #     worksheet.set_column(row, column, len(str(data)))
            # else:
            #     worksheet.set_column(row, column, 8)
            worksheet.set_column(row, column, 10)
            column += 1
        row += 1
        column = 0
        print (order_data['customer_id'])
    
    workbook.close()